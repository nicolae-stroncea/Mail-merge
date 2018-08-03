import openpyxl
import datetime
import os
import send_email as em
import shutil

def send_text_files(file, list_of_dictionaries, fields, subject):
    '''Sends text files through email'''
    clear_directory(fields[0])
    for dictionary in list_of_dictionaries:
        # Create the text file
        out_file = create_text_file(file, dictionary, fields)
        # Get it in string form
        email_body = em.get_email_body(out_file)
        # Send the emails
        em.send_email(dictionary["email"], email_body, subject)


def first_last_file(file, list_of_dictionaries, fields, subject):
    '''Creates a list which contains 2 dictionaries, for the first person, and
    the last last. Each dictionary contains
    the address for an email,the subject, and the body.'''
    clear_directory(fields[0])
    i = 0
    listofEmails = []
    indexes_of_dictionaries = (0, len(list_of_dictionaries) - 1)
    for i in range(2):
        dictionary = list_of_dictionaries[indexes_of_dictionaries[i]]
        # Create the text file
        out_file = create_text_file(file, dictionary, fields)
        # Get it in string form
        email_body = em.get_email_body(out_file)
        # Store the body, the subject, and the email address into dict
        entire_email = {}
        entire_email["subject"] = subject
        entire_email["email"] = dictionary["email"]
        entire_email["email_body"] = email_body
        i += 1
        listofEmails.append(entire_email)
    return listofEmails


def dictList(sheet):  # max_column
    '''(int, int, str) -> list
    Given an integer which'''
    # iterates through every row until the last one
    # Starts from 2 since the first row is the categories
    row = 2
    rowDictList = []
    while(row <= sheet.max_row and sheet.cell(row, 1).value is not None):
        # Create a row dictionary
        rowdict = {}
        # Row number says the number of the row
        rowdict["row_number"] = row
        column = 1
        # For this given row, we will
        # Store the name of the columns as attributes,
        # and the values from those
        # rows as the values of the attributes
        # while column <= max_column:
        while (column <= sheet.max_column and
               sheet.cell(1, column).value is not None):
            # Make the attribute, based on the first row which have the
            # categories
            attribute = sheet.cell(1, column).value
            rowdict[attribute] = sheet.cell(row, column).value
            column += 1
        rowDictList.append(rowdict)
        row += 1
    return rowDictList


def clear_directory(path):
    dir = str(path)
    if not os.path.exists(dir):
        os.makedirs(dir)
    else:
        shutil.rmtree(dir)  # removes all the subdirectories! if exists
        os.makedirs(dir)


def create_text_file(file, dictionary, fields):
    '''(TextIOWrapper, list) -> TextIOWrapper
    Creates a text file.'''
    # Iterate through every dictionary. Will create one new file for each one.
    # Initiate the fields_index
    fields_index = 0
    all_lines = file.readlines()
    new_file_name = str(fields[fields_index]) + "/" + str(
        dictionary["Name"]) + ".txt"
    fields_index += 1
    f = open(new_file_name, "w+")
    for line in all_lines:
        # We can have more than 1 in the same line
        # iterate through the line looking for the three ###
        counter = 0
        # Iterate through line str, character by character
        while counter < len(line):
            if line[counter] == "#":
                if line[counter + 1] == "#" and line[counter + 2] == "#":
                    # Potential Error Create Exception
                    # Format the different possible values according to their
                    # categories
                    attribute = fields[fields_index]
                    if attribute == "Time":
                        # %I gets time in 12 hour format
                        # %p gets AM or PM
                        value = str(dictionary[attribute].strftime("%I")) +\
                            ":" +  str(dictionary[attribute].minute) +\
                            " " + str(dictionary[attribute].strftime("%p"))
                    else:
                        value = str(dictionary[attribute])
                    line = line[:counter] + value + line[counter + 3:]
                    fields_index += 1
                    counter += 3
            else:
                counter += 1
        f.write(line)
    # Reset counter to beginning of the file
    file.seek(0)
    f.close()
    return new_file_name


